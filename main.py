import speech_recognition as sr
import cv2
import face_recognition
import json
import os
import random
import pyttsx3
import time
import openpyxl

class Main:
    # Initialize recognizer class
    data_path = "data/Resources/data"
    isMaster = False
    cap = cv2.VideoCapture(0)

    @staticmethod
    def load():
        encoding_file = os.path.join(Main.data_path, 'encodings.json')
        user_data_file = os.path.join(Main.data_path, 'data.json')
        return encoding_file, user_data_file

    @staticmethod
    def load_user_data(data_file):
        with open(data_file, 'r') as file:
            return json.load(file)

    @staticmethod
    def load_encodings(encoding_file):
        with open(encoding_file, 'r') as file:
            return json.load(file)

    recognizer = sr.Recognizer()
    engine = pyttsx3.init()

    # Setting up speech rate
    rate = engine.getProperty('rate')
    voice = engine.getProperty('voices')
    engine.setProperty('voice', voice[0].id)
    engine.setProperty('rate', rate - 50)  # Reducing speech rate by 50 words per minute

    isCompleted = False
    isHasCode = False
    isFirst = True

    @staticmethod
    def start_recognition(encoding_file, user_data_file):
        Main.isCompleted = False

        while not Main.isMaster:
            ret, frame = Main.cap.read()
            if not ret:
                break

            # Face Recognition
            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            face_locations = face_recognition.face_locations(rgb_frame)
            face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

            # Process face recognition
            for face_location, face_encoding in zip(face_locations, face_encodings):
                top, right, bottom, left = face_location

                # Draw a rectangle around the face
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)

                # Check if the face matches any known encoding
                match = False
                for name, encoding in encodings.items():
                    if face_recognition.compare_faces([encoding], face_encoding)[0]:
                        user_info = user_data.get("User", {}).get(name)
                        if user_info:
                            cv2.putText(frame, f"{user_info['name']}", (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
                            match = True
                            if name == "0001" and Main.isFirst:
                                Main.speak(f'Welcome back Master {user_info["name"]}')
                                Main.isFirst = False
                                Main.isMaster = True
                                break
                            elif Main.isFirst:
                                Main.speak(f'Welcome back {user_info["name"]}')
                                Main.speak(f'{user_info["name"]}! Do you have any permission to use? if you had enjoy!')
                                Main.isFirst = False
                                Main.isMaster = True
                                break
                        else:
                            pass
        
        # Turn off the camera and close OpenCV windows
        Main.cap.release()
        cv2.destroyAllWindows()
        while Main.isMaster:
            with sr.Microphone() as source:
                print("Speak something...")
                Main.recognizer.adjust_for_ambient_noise(source)  # Adjust for ambient noise
                audio = Main.recognizer.listen(source)

            try:
                print("Recognizing...")
                text = Main.recognizer.recognize_google(audio)
                if text is not None:
                    if Main.isHasCode:
                        Main.data_handled(text.lower())

                    if text.lower() == "117" and not Main.isCompleted:
                        print("ヾ(⌐■_■)ノ You:", text)
                        Main.isHasCode = not Main.isHasCode

                    elif text.lower() == "stop":
                        Main.speak("Bye see you later!")
                        print("(づ￣ 3￣)づ:....!")
                        break
                else:
                    print("(⊙_⊙)？")
            except sr.UnknownValueError:
                print("Sorry, could not understand audio.")
            except sr.RequestError as e:
                print("Error fetching results; {0}".format(e))

        Main.cap.release()
        cv2.destroyAllWindows()

    @staticmethod
    def data_excel():
        wb_path = os.path.join(Main.data_path, 'data.xlsx')
        try:
            wb = openpyxl.load_workbook(wb_path)
            wu = wb['User']
            wr = wb['Response']

            greeting_list = [wu.cell(row=1, column=i).value for i in range(1, wu.max_column + 1)]
            reply_greeting_list = [wr.cell(row=1, column=i).value for i in range(1, wr.max_column + 1)]

            return greeting_list, reply_greeting_list

        except FileNotFoundError:
            print("Excel file not found.")
            return None, None
        except Exception as e:
            print("An error occurred:", e)
            return None, None

    @staticmethod
    def data_handled(text):
        greeting_list, reply_greeting_list = Main.data_excel()
        if Main.isHasCode and greeting_list is not None:
            text_lower = text.lower()
            for greeting in greeting_list:
                if greeting.lower() == text_lower:
                    response = random.choice(reply_greeting_list)
                    print("ヾ(⌐■_■)ノ You:", text)
                    print("o(*￣︶￣*)o:", response)
                    Main.speak(response) 
                    time.sleep(2)
                    return
            print("(￣﹃￣)...?:", text)



    @staticmethod
    def speak(text):
        Main.engine.say(text)
        Main.engine.runAndWait()

if __name__ == '__main__':
    # Load encoding and user data from JSON files
    encoding_file_path, user_data_file_path = Main.load()
    with open(encoding_file_path, 'r') as encoding_file:
        encodings = json.load(encoding_file)
    with open(user_data_file_path, 'r') as user_data_file:
        user_data = json.load(user_data_file)
    # Pass the loaded data to start_recognition method
    Main.start_recognition(encodings, user_data)
